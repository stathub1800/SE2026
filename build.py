# -*- coding: utf-8 -*-
"""
Pembangun halaman Struktur Progres SE2026 - BPS Provinsi Lampung.

Cara pakai:
    1. Taruh dua berkas ekspor FASIH terbaru di folder  data\\
    2. Jalankan:  python build.py
    3. Hasilnya:  index.html  (siap di-commit ke GitHub Pages)

Kolom dicari berdasarkan JUDUL kolom, bukan nomor urutnya, sehingga tetap
bekerja walau FASIH menyisipkan kolom baru di tengah tabel.

Nama kartu, definisi, dan rumus ada di daftar N di bawah. Ubah "label" untuk
mengganti nama kartu; jangan ubah "id" karena itu kunci penghubung antar kartu.
"""
import sys, json, datetime, pathlib, re
import pandas as pd
from openpyxl import load_workbook

AKAR = pathlib.Path(__file__).resolve().parent
DATA = AKAR / "data"

def cari_berkas(pola, keterangan):
    hit = sorted(DATA.glob(pola))
    if not hit:
        sys.exit(f"GAGAL: berkas {keterangan} tidak ditemukan di {DATA}\n"
                 f"       Dicari dengan pola: {pola}")
    return max(hit, key=lambda p: p.stat().st_mtime)

F1 = cari_berkas("*Pendataan*.xlsx", "Progres Pendataan")
F2 = cari_berkas("*Pemutakhiran*Keluarga*.xlsx", "Pemutakhiran Keluarga")
print(f"  Pendataan   : {F1.name}")
print(f"  Pemutakhiran: {F2.name}")

# ---------------------------------------------------------------- pemetaan kolom
def bersih(t):
    return re.sub(r"\s+", " ", str(t or "").replace("\u200b", "").replace("\xa0", " ")).strip()

def struktur(path, sheet):
    """Kembalikan (daftar judul tiap kolom, nomor baris data pertama)."""
    ws = load_workbook(path, data_only=True)[sheet]
    A = lambda r: bersih(ws.cell(row=r, column=1).value)
    try:
        baris_data = next(r for r in range(1, 40) if A(r).isdigit())
        baris_kode = next(r for r in range(1, baris_data) if A(r).lower() == "kode")
    except StopIteration:
        sys.exit(f"GAGAL: susunan judul lembar '{sheet}' tidak dikenali.")
    judul = list(range(baris_kode, baris_data - 1))
    isi = {(r, c): bersih(ws.cell(row=r, column=c).value)
           for r in judul for c in range(1, ws.max_column + 1)}
    for rng in ws.merged_cells.ranges:
        v = bersih(ws.cell(row=rng.min_row, column=rng.min_col).value)
        for r in range(rng.min_row, rng.max_row + 1):
            if r in judul:
                for c in range(rng.min_col, rng.max_col + 1):
                    isi[(r, c)] = v
    label = []
    for c in range(1, ws.max_column + 1):
        p = []
        for r in judul:
            t = isi.get((r, c), "")
            if t and (not p or p[-1] != t):
                p.append(t)
        label.append(" > ".join(p))
    return label, baris_data - 1

class Lembar:
    def __init__(self, path, sheet):
        self.nama = sheet
        self.label, skip = struktur(path, sheet)
        d = pd.read_excel(path, sheet_name=sheet, header=None, skiprows=skip)
        d = d[pd.to_numeric(d[0], errors="coerce").notna()].copy()
        d[0] = d[0].astype(int).astype(str)
        self.df = d.set_index(0)
        self.peta = {}

    def kolom(self, nama, ada=(), tanpa=(), daun=None, wajib=True):
        hit = []
        for i, lab in enumerate(self.label):
            l = lab.lower()
            if any(a.lower() not in l for a in ada):        continue
            if any(t.lower() in l for t in tanpa):          continue
            if daun is not None and bersih(lab.split(" > ")[-1]).lower() != daun.lower(): continue
            hit.append(i)
        if len(hit) != 1:
            if not wajib:
                return None
            sisa = "tidak ada" if not hit else "lebih dari satu: " + ", ".join(
                f"[{i}] {self.label[i][:60]}" for i in hit)
            sys.exit(f"GAGAL: kolom '{nama}' pada lembar '{self.nama}' {sisa}.\n"
                     f"       Kemungkinan susunan ekspor FASIH berubah lagi.\n"
                     f"       Judul kolom yang terbaca:\n" +
                     "\n".join(f"         [{i}] {x}" for i, x in enumerate(self.label)))
        self.peta[nama] = hit[0]
        return hit[0]

    def nilai(self, kode, kol):
        if kol is None:
            return 0
        v = self.df.loc[kode, kol]
        return 0 if pd.isna(v) else int(v)

    def persen(self, kode, kol):
        if kol is None:
            return None
        v = self.df.loc[kode, kol]
        if pd.isna(v):
            return None
        return float(str(v).replace(".", "").replace(",", ".")) if isinstance(v, str) else float(v)

pp  = Lembar(F1, "PROGRES PENDATAAN")
up  = Lembar(F1, "USAHA PERUSAHAAN")
sk  = Lembar(F1, "SKALA USAHA")
uk  = Lembar(F1, "USAHA KELUARGA")
ku  = Lembar(F1, "KESELURUHAN USAHA")
kel = Lembar(F2, "KELUARGA")
ak  = Lembar(F2, "ANGGOTA KELUARGA")
kk  = Lembar(F2, "KELUARGA KHUSUS")

K = {}
TAMBAHAN = {}
K["pre_assign"]  = pp.kolom("prelist assignment", ["jumlah assignment"], daun="Prelist")
K["baru_assign"] = pp.kolom("assignment baru",    ["jumlah assignment"], daun="Baru")
K["verif"]       = pp.kolom("hasil verifikasi",   ["hasil verifikasi"], tanpa=["persentase"])
K["p_verif"]     = pp.kolom("% hasil verifikasi", ["persentase hasil verifikasi"])
K["responden"]   = pp.kolom("responden didata",   ["responden didata"], tanpa=["persentase", "sedang"])
K["p_responden"] = pp.kolom("% responden didata", ["persentase responden didata"], tanpa=["sedang"])
K["draft"]       = pp.kolom("draft",              ["sedang didata"], tanpa=["persentase"])
K["p_draft"]     = pp.kolom("% draft",            ["persentase", "sedang didata"])
# kolom Open baru muncul pada ekspor Agustus 2026; kalau tidak ada, Open dihitung sebagai sisa
K["open_ub"]  = pp.kolom("open UB",       ["belum didata"], daun="UB",       wajib=False)
K["open_um"]  = pp.kolom("open UM",       ["belum didata"], daun="UM",       wajib=False)
K["open_umk"] = pp.kolom("open UMK",      ["belum didata"], daun="UMK",      wajib=False)
K["open_kel"] = pp.kolom("open keluarga", ["belum didata"], daun="Keluarga", wajib=False)
ADA_OPEN = all(K[x] is not None for x in ("open_ub", "open_um", "open_umk", "open_kel"))

K["pre_usaha"] = up.kolom("prelist usaha", ["jumlah prelist usaha"], tanpa=["keluarga"])
K["force"]     = up.kolom("force submit",  ["force submit"], tanpa=["persentase"])
K["p_force"]   = up.kolom("% force submit",["persentase force submit"])
K["bku"]       = up.kolom("usaha BKU",     ["usaha bku ("], tanpa=["persentase"])
K["p_bku"]     = up.kolom("% usaha BKU",   ["persentase", "usaha bku ("])

SKALA = "menurut status skala"
K["pre_ub"]  = sk.kolom("prelist UB",  ["jumlah prelist"], daun="UB")
K["pre_um"]  = sk.kolom("prelist UM",  ["jumlah prelist"], daun="UM")
K["pre_umk"] = sk.kolom("prelist UMK", ["jumlah prelist"], daun="UMK")
K["ub"]      = sk.kolom("UB tercacah",  [SKALA], daun="UB")
K["um"]      = sk.kolom("UM tercacah",  [SKALA], daun="UM")
K["umk"]     = sk.kolom("UMK tercacah", [SKALA], daun="UMK")
K["p_ub"]    = sk.kolom("% UB",  [SKALA], daun="Persentase UB")
K["p_um"]    = sk.kolom("% UM",  [SKALA], daun="Persentase UM")
K["p_umk"]   = sk.kolom("% UMK", [SKALA], daun="Persentase UMK")
K["unk"]     = sk.kolom("skala tak terklasifikasi", ["tidak dapat diklasifikasikan"])

UKG = "usaha keluarga menurut status"
K["pre_uk"]  = uk.kolom("prelist usaha keluarga", ["jumlah prelist usaha keluarga"])
K["uk_ok"]   = uk.kolom("usaha keluarga ditemukan", [UKG], daun="Ditemukan")
K["uk_baru"] = uk.kolom("usaha keluarga baru",      [UKG], daun="Baru")
K["uk"]      = uk.kolom("usaha dalam keluarga",   ["jumlah usaha dalam keluarga"])
K["p_uk"]    = uk.kolom("% usaha dalam keluarga", ["persentase usaha dalam keluarga"])

K["usaha_all"]     = ku.kolom("total usaha",         daun="Total Usaha")
K["p_usaha_all"]   = ku.kolom("% total usaha",       daun="Persentase Total Usaha")
K["pre_usaha_all"] = ku.kolom("total prelist usaha", ["total prelist usaha dan usaha keluarga"])

K["pre_kel"]  = kel.kolom("prelist awal keluarga", ["prelist awal"])
K["kel_ok"]   = kel.kolom("keluarga ditemukan",    daun="Ditemukan")
K["p_kelok"]  = kel.kolom("% keluarga ditemukan",  daun="Persentase Ditemukan")
K["kel_baru"] = kel.kolom("keluarga baru",         ["keluarga baru"])
K["kel"]      = kel.kolom("total hasil keluarga",  ["total hasil pendataan"], tanpa=["persentase"])
K["p_kel"]    = kel.kolom("% total hasil keluarga",["persentase total hasil pendataan"])

K["ak_bersama"] = ak.kolom("AK tinggal bersama", ["tinggal bersama"])
K["ak_baru"]    = ak.kolom("AK baru",            ["anggota keluarga baru"])
K["ak"]         = ak.kolom("total AK",           ["total anggota keluarga"])

K["kk_bangunan"] = kk.kolom("bangunan K1 hasil listing", ["hasil pendataan ppl"])
K["kk_didata"]   = kk.kolom("bangunan K1 didata",  ["khusus didata"], tanpa=["persentase"])
K["p_kk"]        = kk.kolom("% bangunan K1 didata",["persentase bangunan keluarga khusus didata"])

# ---- kolom rincian tambahan, untuk memperkaya tabel di popup ----
UBG, UMG = "> ub >", "> umkm >"
for nm, lem, ada, daun in [
    ("ub_ok",    up, [UBG], "Ditemukan/Pindah Dapat Ditelusuri"),
    ("ub_bukan", up, [UBG], "Ditemukan Tetapi Bukan Cakupan Survei"),
    ("ub_pindah",up, [UBG], "Pindah Tidak Dapat Ditelusuri"),
    ("ub_tutup", up, [UBG], "Tutup"),
    ("ub_dup",   up, [UBG], "Duplikat"),
    ("ub_tt",    up, [UBG], "Tidak Ditemukan"),
    ("ub_kp",    up, [UBG], "Data Diperoleh dari Kantor Pusat"),
    ("ub_nr",    up, [UBG], "Nonrespon"),
    ("ub_total", up, [UBG], "Total UB"),
    ("mk_ok",    up, [UMG], "Ditemukan"),
    ("mk_tutup", up, [UMG], "Tutup"),
    ("mk_ganda", up, [UMG], "Ganda"),
    ("mk_tt",    up, [UMG], "Tidak Ditemukan"),
    ("mk_baru",  up, [UMG], "Baru"),
    ("mk_kp",    up, [UMG], "Data Diperoleh dari Kantor Pusat"),
    ("mk_nr",    up, [UMG], "Nonrespon"),
    ("mk_total", up, [UMG], "Total UMKM"),
    ("kel_meninggal", kel, [], "Meninggal"),
    ("kel_tdkelig",   kel, [], "Tidak Eligible"),
    ("kel_tt",        kel, [], "Tidak Ditemukan"),
    ("kel_nr",        kel, [], "Nonrespon"),
    ("ak_meninggal",  ak,  [], "Meninggal"),
    ("ak_dn",         ak,  [], "Pindah Dalam Negeri (DN)"),
    ("ak_ln",         ak,  [], "Pindah Luar Negeri (LN)"),
    ("ak_tt",         ak,  [], "Tidak Ditemukan"),
    ("ak_khusus",     ak,  [], "Anggota Keluarga Khusus"),
    ("uk_tutup", uk, [], "Tutup"),
    ("uk_ganda", uk, [], "Ganda"),
    ("uk_tt",    uk, [], "Tidak Ditemukan"),
    ("uk_nr",    uk, [], "Nonrespon"),
]:
    K[nm] = lem.kolom(nm, ada, daun=daun, wajib=False)
    TAMBAHAN[nm] = lem
K["bku_up"] = K["bku"]                      # total BKU versi lembar USAHA/PERUSAHAAN
K["bku"]    = sk.kolom("total BKU (skala)",   ["menurut status skala"], daun="Total Usaha BKU")
K["p_bku"]  = sk.kolom("% total BKU (skala)", ["menurut status skala"], daun="Persentase Total Usaha BKU")

hilang = [n for n, i in K.items() if i is None and n not in ("open_ub","open_um","open_umk","open_kel")]
if hilang:
    print("  CATATAN kolom rincian tidak ditemukan (kolomnya akan berisi 0): " + ", ".join(hilang))

print("\nPeta kolom yang terbaca:")
for lb, obj in (("PROGRES PENDATAAN", pp), ("USAHA PERUSAHAAN", up), ("SKALA USAHA", sk),
                ("USAHA KELUARGA", uk), ("KESELURUHAN USAHA", ku), ("KELUARGA", kel),
                ("ANGGOTA KELUARGA", ak), ("KELUARGA KHUSUS", kk)):
    isi = ", ".join(f"{n}={i}" for n, i in obj.peta.items())
    print(f"  {lb:<20} {isi}")
print("  kolom Open resmi: " + ("ADA, dipakai langsung" if ADA_OPEN else "tidak ada, Open dihitung sebagai sisa"))

# ---------------------------------------------------------------- rakit data
kodes_semua = list(pp.df.index)
rows = {}
for k in kodes_semua:
    r = {"nama": str(pp.df.loc[k, 1])}
    r["pre_assign"] = pp.nilai(k, K["pre_assign"]); r["baru_assign"] = pp.nilai(k, K["baru_assign"])
    r["tot_assign"] = r["pre_assign"] + r["baru_assign"]
    r["verif"] = pp.nilai(k, K["verif"]); r["responden"] = pp.nilai(k, K["responden"])
    r["draft"] = pp.nilai(k, K["draft"])
    if ADA_OPEN:
        r["open_ub"]  = pp.nilai(k, K["open_ub"]);  r["open_um"]  = pp.nilai(k, K["open_um"])
        r["open_umk"] = pp.nilai(k, K["open_umk"]); r["open_kel"] = pp.nilai(k, K["open_kel"])
        r["open"] = r["open_ub"] + r["open_um"] + r["open_umk"] + r["open_kel"]
    else:
        r["open"] = max(r["tot_assign"] - r["verif"] - r["draft"], 0)
    r["nihil"] = r["verif"] - r["responden"]
    r["pre_usaha"] = up.nilai(k, K["pre_usaha"]); r["bku"] = sk.nilai(k, K["bku"])
    r["force"] = up.nilai(k, K["force"]); r["usaha_unit"] = r["bku"] + r["force"]
    for f in ("ub", "um", "umk", "unk", "pre_ub", "pre_um", "pre_umk"):
        r[f] = sk.nilai(k, K[f])
    for f in ("pre_uk", "uk_ok", "uk_baru", "uk"):
        r[f] = uk.nilai(k, K[f])
    r["usaha_all"] = ku.nilai(k, K["usaha_all"]); r["pre_usaha_all"] = ku.nilai(k, K["pre_usaha_all"])
    for f in ("pre_kel", "kel_ok", "kel_baru", "kel"):
        r[f] = kel.nilai(k, K[f])
    for f in ("ak_bersama", "ak_baru", "ak"):
        r[f] = ak.nilai(k, K[f])
    r["kk_bangunan"] = kk.nilai(k, K["kk_bangunan"]); r["kk_didata"] = kk.nilai(k, K["kk_didata"])
    r["p_verif"] = pp.persen(k, K["p_verif"]); r["p_responden"] = pp.persen(k, K["p_responden"])
    r["p_draft"] = pp.persen(k, K["p_draft"])
    r["p_bku"] = sk.persen(k, K["p_bku"]); r["p_force"] = up.persen(k, K["p_force"])
    r["p_ub"] = sk.persen(k, K["p_ub"]); r["p_um"] = sk.persen(k, K["p_um"]); r["p_umk"] = sk.persen(k, K["p_umk"])
    r["p_uk"] = uk.persen(k, K["p_uk"]); r["p_usaha_all"] = ku.persen(k, K["p_usaha_all"])
    r["p_kel"] = kel.persen(k, K["p_kel"]); r["p_kelok"] = kel.persen(k, K["p_kelok"])
    r["p_kk"] = kk.persen(k, K["p_kk"])
    r["p_open"] = round(r["open"] / r["pre_assign"] * 100, 2) if r["pre_assign"] else None
    r["bku_up"] = up.nilai(k, K["bku_up"])
    for nm, lem in TAMBAHAN.items():
        r[nm] = lem.nilai(k, K[nm])
    r["umk_bintang"] = r["umk"] + r["unk"]              # UMK* seperti tampilan FASIH
    r["usaha_all"] = r["bku"] + r["uk"]                 # dihitung dari sumber tersegar
    r["p_usaha_all"] = round(r["usaha_all"] / r["pre_usaha_all"] * 100, 2) if r["pre_usaha_all"] else None
    r["selisih_bku"] = r["bku"] - r["bku_up"]
    rows[k] = r

# ---------------------------------------------------------------- uji mutu
def uji(nama, fn):
    salah = [k for k in kodes_semua if fn(rows[k])[0] != fn(rows[k])[1]]
    print(("  OK    " if not salah else "  GAGAL ") + nama + ("" if not salah else "  -> " + ", ".join(salah)))
    return not salah

print("\nUji identitas penjumlahan:")
lulus = all([
    uji("prelist keluarga + prelist usaha = prelist assignment", lambda r: (r["pre_kel"]+r["pre_usaha"], r["pre_assign"])),
    uji("responden + hasil nihil = verifikasi",                  lambda r: (r["responden"]+r["nihil"], r["verif"])),
    uji("keluarga + usaha BKU + force submit = responden",       lambda r: (r["kel"]+r["bku_up"]+r["force"], r["responden"])),
    uji("UB + UM + UMK + lainnya = usaha BKU",                   lambda r: (r["ub"]+r["um"]+r["umk"]+r["unk"], r["bku"])),
    uji("keluarga ditemukan + baru = keluarga",                  lambda r: (r["kel_ok"]+r["kel_baru"], r["kel"])),
    uji("verifikasi + draft + open = total assignment",          lambda r: (r["verif"]+r["draft"]+r["open"], r["tot_assign"])),
])

gap = [k for k in kodes_semua if rows[k]["selisih_bku"] != 0]
CATATAN_EKSPOR = ""
if gap:
    s = rows["18"]["selisih_bku"]
    CATATAN_EKSPOR = (f"Total Usaha BKU berbeda antar lembar pada ekspor ini: lembar Skala Usaha "
                      f"{rows['18']['bku']:,} sedangkan lembar Usaha/Perusahaan {rows['18']['bku_up']:,}, "
                      f"selisih {abs(s):,}. Halaman ini memakai angka lembar Skala Usaha karena itu yang "
                      f"tampil di layar FASIH. Akibatnya jumlah Usaha ditambah Keluarga melebihi Responden "
                      f"Didata sebesar selisih tersebut. Biasanya karena kedua lembar ditarik pada waktu "
                      f"berbeda dan akan lurus kembali pada tarikan berikutnya.").replace(",", ".")
    print(f"  CATATAN Total Usaha BKU beda antar lembar di {len(gap)} wilayah "
          f"(Lampung: skala {rows['18']['bku']:,} vs usaha/perusahaan {rows['18']['bku_up']:,}).")
    print("          Halaman memakai angka lembar Skala Usaha, sesuai tampilan layar FASIH.")

neg = [f"{rows[k]['nama']} ({rows[k]['open']:,})".replace(",", ".") for k in kodes_semua if rows[k]["open"] < 0]
if neg:
    print("  CATATAN Open bernilai negatif di: " + "; ".join(neg))
    print("          Angka ini apa adanya dari FASIH. Terjadi bila assignment yang sudah dikerjakan")
    print("          melampaui prelist wilayah tersebut. Jelaskan bila ditanya pimpinan.")

print("\nUji persentase resmi FASIH terhadap hitung ulang:")
pasangan = [("p_verif","verif","pre_assign"),("p_responden","responden","pre_assign"),("p_draft","draft","pre_assign"),
            ("p_bku","bku","pre_usaha"),("p_force","force","pre_usaha"),("p_ub","ub","pre_ub"),("p_um","um","pre_um"),
            ("p_umk","umk","pre_umk"),("p_uk","uk","pre_uk"),("p_kel","kel","pre_kel"),("p_kelok","kel_ok","pre_kel"),
            ("p_kk","kk_didata","kk_bangunan"),("p_usaha_all","usaha_all","pre_usaha_all")]
beda = 0
for pf, vf, bf in pasangan:
    for k in kodes_semua:
        r = rows[k]
        if r[pf] is None or not r[bf]:
            continue
        if abs(round(r[vf]/r[bf]*100, 2) - r[pf]) > 0.02:
            beda += 1
            print(f"  selisih {k} {pf}: berkas {r[pf]} vs hitung {round(r[vf]/r[bf]*100,2)}")
print(f"  {'OK    semua cocok' if beda == 0 else f'PERIKSA {beda} selisih'}")

kodes = sorted(k for k in rows if k != "18")

N=[
 dict(id='assign',parent=None,label='Assignment SE2026',field='tot_assign',base=None,unit='assignment',tipe='akar',warna='maroon',
   ket='Seluruh tugas pencacahan di FASIH: prelist berjalan ditambah assignment baru dari lapangan. Prelist berjalan adalah jumlah prelist terkini di FASIH, bersifat dinamis mengikuti perpindahan wilayah assignment, sehingga angkanya dapat bergeser sedikit dari hari ke hari.',
   rumus='prelist berjalan + baru',parts=[('Prelist berjalan','pre_assign'),('Baru dari lapangan','baru_assign')],
   sumber='Progres Pendataan kol. (3) + (4)',turunan='Penjumlahan dua kolom. Tidak ada satu kolom tunggal di ekspor yang memuat angka ini.'),

 dict(id='capaian',parent='assign',label='Verifikasi',sub='Hasil verifikasi lapangan',field='verif',base='tot_assign',unit='assignment',tipe='bagi',warna='hijau',
   ket='Jumlah hasil verifikasi lapangan assignment usaha dan keluarga yang tidak berstatus Open atau Draft. Mencakup seluruh status keberadaan: ditemukan, tutup, ganda, tidak ditemukan, baru, dan lainnya.',
   rumus='hasil verifikasi \u00f7 total assignment',sumber='Progres Pendataan kol. (5)',
   resmi=dict(p='p_verif',b='pre_assign',bl='prelist assignment',kol='Progres Pendataan kol. (6)')),
 dict(id='draft',parent='assign',label='Draft',field='draft',base='tot_assign',unit='assignment',tipe='bagi',warna='kuning',neg=1,
   ket='Assignment yang sudah dibuka dan sedang diisi tetapi belum disubmit. Pekerjaan menggantung yang perlu didorong tuntas.',
   rumus='responden sedang didata \u00f7 total assignment',sumber='Progres Pendataan kol. (9)',
   resmi=dict(p='p_draft',b='pre_assign',bl='prelist assignment',kol='Progres Pendataan kol. (10)')),
 dict(id='open',parent='assign',label='Open',field='open',base='tot_assign',unit='assignment',tipe='bagi',warna='merah',neg=1,
   ket='Assignment yang belum disentuh sama sekali. Sisa beban murni yang masih harus dikerjakan.',
   rumus='total assignment \u2212 verifikasi \u2212 draft',sumber='hitungan turunan',
   turunan='Ekspor FASIH tidak memuat kolom Open. Angka ini sisa dari total assignment setelah dikurangi verifikasi dan draft.'),

 dict(id='responden',parent='capaian',label='Responden Didata',sub='Responden hasil pendataan',field='responden',base='verif',unit='responden',tipe='bagi',warna='hijau',
   ket='Jumlah responden usaha (BKU) dan keluarga yang telah berhasil didata, yaitu selain berstatus Open atau Draft, dengan kriteria keberadaan ditemukan, baru, atau force submit.',
   rumus='responden didata \u00f7 verifikasi',sumber='Progres Pendataan kol. (7)',
   resmi=dict(p='p_responden',b='pre_assign',bl='prelist assignment',kol='Progres Pendataan kol. (8)')),
 dict(id='nihil',parent='capaian',label='Hasil Nihil',field='nihil',base='verif',unit='assignment',tipe='bagi',warna='abu',neg=1,
   ket='Assignment yang sudah terverifikasi tetapi status keberadaannya di luar kriteria responden hasil pendataan, yaitu bukan ditemukan, baru, atau force submit. Termasuk usaha tutup, ganda, pindah tak tertelusuri, tidak ditemukan, serta keluarga meninggal dan tidak eligible. Pekerjaannya tetap dihitung selesai.',
   rumus='verifikasi \u2212 responden didata',sumber='hitungan turunan',
   turunan='Ekspor FASIH tidak memuat kolom ini. Angka ini selisih antara hasil verifikasi dan responden didata.'),

 dict(id='usaha',parent='responden',label='Usaha',field='usaha_unit',base='responden',unit='usaha',tipe='bagi',warna='biru',
   ket='Responden berupa unit usaha dari assignment usaha, mencakup usaha BKU yang berstatus keberadaan ditemukan atau baru, ditambah assignment usaha yang disubmit paksa karena ternyata berskala usaha keluarga (force submit).',
   rumus='usaha BKU + force submit',parts=[('Usaha BKU','bku'),('Force submit','force')],
   sumber='Usaha/Perusahaan kol. (35) + (36)',
   turunan='Penjumlahan dua kolom status. Ekspor tidak memuat gabungan ini dalam satu kolom.',
   resmi=dict(p='p_bku',f='bku',fl='Usaha BKU',b='pre_usaha',bl='prelist usaha',kol='Usaha/Perusahaan kol. (37)')),
 dict(id='keluarga',parent='responden',label='Keluarga',field='kel',base='responden',unit='keluarga',tipe='bagi',warna='oranye',
   ket='Total keluarga hasil pendataan, yaitu jumlah keluarga ditemukan ditambah keluarga baru.',
   rumus='keluarga ditemukan + keluarga baru',parts=[('Ditemukan','kel_ok'),('Baru','kel_baru')],
   sumber='Pemutakhiran Keluarga kol. (15)',
   resmi=dict(p='p_kel',b='pre_kel',bl='prelist awal keluarga',kol='Keluarga kol. (16)')),

 dict(id='ub',parent='usaha',label='UB',sub='Usaha Besar',field='ub',base='usaha_unit',unit='usaha',tipe='bagi',warna='maroon',
   ket='Usaha Besar yang berhasil didata dari responden berstatus keberadaan ditemukan atau baru. Jumlahnya bisa melampaui prelist karena ada usaha yang naik kelas setelah dicacah.',
   rumus='UB tercacah \u00f7 usaha',sumber='Skala Usaha kol. (7)',
   resmi=dict(p='p_ub',b='pre_ub',bl='prelist UB',kol='Skala Usaha kol. (8)')),
 dict(id='um',parent='usaha',label='UM',sub='Usaha Menengah',field='um',base='usaha_unit',unit='usaha',tipe='bagi',warna='biru',
   ket='Usaha Menengah yang berhasil didata dari responden berstatus keberadaan ditemukan atau baru.',rumus='UM tercacah \u00f7 usaha',sumber='Skala Usaha kol. (9)',
   resmi=dict(p='p_um',b='pre_um',bl='prelist UM',kol='Skala Usaha kol. (10)')),
 dict(id='umk',parent='usaha',label='UMK',sub='Usaha Mikro Kecil',field='umk',base='usaha_unit',unit='usaha',tipe='bagi',warna='biru',
   ket='Usaha Mikro dan Kecil yang berhasil didata dari responden berstatus keberadaan ditemukan atau baru. Di layar FASIH angka ini ditulis UMK* karena sudah digabung dengan unit pembantu atau penunjang; pada halaman ini keduanya sengaja dipisah agar penjumlahannya terlihat. Lihat kolom UMK* pada tabel di bawah untuk angka gabungannya.',
   rumus='UMK tercacah \u00f7 usaha',sumber='Skala Usaha kol. (11)',
   resmi=dict(p='p_umk',b='pre_umk',bl='prelist UMK',kol='Skala Usaha kol. (12)')),
 dict(id='unk',parent='usaha',label='Unit Pembantu',sub='Penunjang, pendapatan nol',field='unk',base='usaha_unit',unit='usaha',tipe='bagi',warna='abu',
   ket='Unit pembantu atau penunjang dengan nilai pendapatan nol, sehingga skala usahanya tidak dapat diklasifikasikan. Ini kondisi wajar, bukan isian yang kurang lengkap. FASIH menggabungkannya dengan UMK dan menuliskannya sebagai UMK*.',
   rumus='unit pembantu (tidak dapat diklasifikasikan) \u00f7 usaha',sumber='Skala Usaha kol. (13)',
   turunan='Ekspor memuat jumlahnya tetapi tidak memuat kolom persentase untuk rincian ini.'),
 dict(id='force',parent='usaha',label='Force Submit',sub='Diklaim sebagai usaha keluarga',field='force',base='usaha_unit',unit='usaha',tipe='bagi',warna='kuning',
   ket='Assignment usaha yang setelah dikunjungi ternyata berskala usaha keluarga, lalu disubmit paksa dan dialihkan ke jalur keluarga. Berbeda dengan kartu Usaha dalam Keluarga yang merupakan roster di dalam assignment keluarga.',
   rumus='force submit \u00f7 usaha',sumber='Usaha/Perusahaan kol. (34)',
   resmi=dict(p='p_force',b='pre_usaha',bl='prelist usaha',kol='Usaha/Perusahaan kol. (35)')),
 dict(id='uk',parent='usaha',label='Usaha dalam Keluarga',field='uk',base=None,unit='usaha',tipe='tautan',warna='biru',
   ket='Usaha keluarga yang berhasil didata dari responden berstatus keberadaan ditemukan atau baru, dicatat sebagai roster di dalam assignment keluarga dan bukan assignment usaha tersendiri. Angka ini tidak menambah jumlah assignment, tetapi menambah jumlah usaha.',
   rumus='usaha dalam keluarga ditemukan + baru',parts=[('Ditemukan','uk_ok'),('Baru','uk_baru')],
   tautan_ket='Tercatat di dalam assignment keluarga',sumber='Usaha Keluarga kol. (16)',
   resmi=dict(p='p_uk',b='pre_uk',bl='prelist usaha keluarga',kol='Usaha Keluarga kol. (17)')),

 dict(id='kelok',parent='keluarga',label='Keluarga Ditemukan',field='kel_ok',base='kel',unit='keluarga',tipe='bagi',warna='oranye',
   ket='Keluarga prelist yang berhasil ditemukan dan dimutakhirkan di lapangan.',
   rumus='ditemukan \u00f7 keluarga',sumber='Pemutakhiran Keluarga kol. (4)',
   resmi=dict(p='p_kelok',b='pre_kel',bl='prelist awal keluarga',kol='Keluarga kol. (5)')),
 dict(id='kelbaru',parent='keluarga',label='Keluarga Baru',field='kel_baru',base='kel',unit='keluarga',tipe='bagi',warna='kuning',
   ket='Keluarga yang belum ada di prelist dan ditambahkan petugas saat pemutakhiran.',
   rumus='keluarga baru \u00f7 keluarga',sumber='Pemutakhiran Keluarga kol. (6)',
   turunan='Ekspor memuat jumlahnya tetapi tidak memuat kolom persentase untuk rincian ini.'),
 dict(id='ak',parent='keluarga',label='Anggota Keluarga',field='ak',base=None,unit='jiwa',tipe='tautan',warna='oranye',
   ket='Jumlah jiwa di dalam keluarga yang terdata. Satuannya orang, bukan keluarga, sehingga wajar jauh lebih besar daripada jumlah keluarga.',
   rumus='tinggal bersama + anggota baru',parts=[('Tinggal bersama','ak_bersama'),('Anggota baru','ak_baru')],
   tautan_ket='Isi dari setiap keluarga',rasio=('kel','jiwa per keluarga'),sumber='Anggota Keluarga kol. (10)',
   turunan='Ekspor tidak memuat kolom persentase untuk anggota keluarga.'),
 dict(id='kkhusus',parent='keluarga',label='Keluarga Khusus',field='kk_didata',base=None,unit='bangunan',tipe='tautan',warna='merah',
   ket='Keluarga yang tinggal di barak, pondok pesantren, hunian sementara, pengungsian, dan sejenisnya. Didata terpisah lewat dokumen K1, sehingga satuan yang dilaporkan di sini adalah bangunan, bukan keluarga.',
   rumus='bangunan didata \u00f7 bangunan hasil pendataan PPL',tautan_ket='Jalur K1, satuan bangunan',
   sumber='Keluarga Khusus kol. (4)',
   resmi=dict(p='p_kk',b='kk_bangunan',bl='bangunan hasil listing PPL',kol='Keluarga Khusus kol. (5)')),
]

# ---- kolom rincian yang ditampilkan pada tabel di setiap popup ----
EKSTRA = {
 "assign":   [("Verifikasi","verif"),("Draft","draft"),("Open","open")],
 "capaian":  [("Responden Didata","responden"),("Hasil Nihil","nihil"),("Draft","draft"),("Open","open")],
 "draft":    [("Verifikasi","verif"),("Open","open")],
 "open":     [("Verifikasi","verif"),("Draft","draft")],
 "responden":[("Keluarga","kel"),("Usaha BKU","bku"),("Force submit","force")],
 "nihil":    [("Klg meninggal","kel_meninggal"),("Klg tdk eligible","kel_tdkelig"),
              ("Klg tdk ditemukan","kel_tt"),("Klg nonrespon","kel_nr"),
              ("Usaha tutup","mk_tutup"),("Usaha ganda","mk_ganda"),("Usaha tdk ditemukan","mk_tt")],
 "usaha":    [("UB","ub"),("UM","um"),("UMK","umk"),("Unit pembantu","unk")],
 "keluarga": [("Meninggal","kel_meninggal"),("Tidak eligible","kel_tdkelig"),
              ("Tidak ditemukan","kel_tt"),("Nonrespon","kel_nr")],
 "ub":       [("Ditemukan","ub_ok"),("Bukan cakupan","ub_bukan"),("Pindah tak tertelusuri","ub_pindah"),
              ("Tutup","ub_tutup"),("Duplikat","ub_dup"),("Tidak ditemukan","ub_tt"),
              ("Dari kantor pusat","ub_kp"),("Nonrespon","ub_nr")],
 "um":       [("UMK","umk"),("Unit pembantu","unk"),("Usaha BKU","bku")],
 "umk":      [("Unit pembantu","unk"),("UMK* (UMK + unit pembantu)","umk_bintang"),
              ("UMKM ditemukan","mk_ok"),("UMKM baru","mk_baru"),("UMKM tutup","mk_tutup"),
              ("UMKM ganda","mk_ganda"),("UMKM tidak ditemukan","mk_tt")],
 "unk":      [("UMK","umk"),("UMK* (UMK + unit pembantu)","umk_bintang")],
 "force":    [("Usaha BKU","bku"),("UMKM ditemukan","mk_ok"),("UMKM baru","mk_baru")],
 "uk":       [("Ditemukan","uk_ok"),("Baru","uk_baru"),("Tutup","uk_tutup"),("Ganda","uk_ganda"),
              ("Tidak ditemukan","uk_tt"),("Nonrespon","uk_nr"),
              ("Total usaha (BKU + dlm keluarga)","usaha_all")],
 "kelok":    [("Meninggal","kel_meninggal"),("Tidak eligible","kel_tdkelig"),
              ("Tidak ditemukan","kel_tt"),("Nonrespon","kel_nr")],
 "kelbaru":  [("Ditemukan","kel_ok"),("Anggota keluarga","ak")],
 "ak":       [("Meninggal","ak_meninggal"),("Pindah dalam negeri","ak_dn"),("Pindah luar negeri","ak_ln"),
              ("Tidak ditemukan","ak_tt"),("AK keluarga khusus","ak_khusus")],
 "kkhusus":  [("AK keluarga khusus","ak_khusus")],
}
for n in N:
    e = [(lb, f) for lb, f in EKSTRA.get(n["id"], []) if any(f in r for r in rows.values())]
    if e:
        n["extra"] = e

# Ekspor Agustus 2026 sudah memuat kolom Open beserta rinciannya.
if ADA_OPEN:
    for n in N:
        if n["id"] == "open":
            n["ket"] = ("Assignment yang belum disentuh sama sekali. Sisa beban murni yang masih harus "
                        "dikerjakan. Sejak ekspor Agustus 2026 angkanya tersedia langsung di FASIH, "
                        "dirinci menurut UB, UM, UMK, dan keluarga.")
            n["rumus"] = "Open UB + UM + UMK + Keluarga"
            n["parts"] = [("UB", "open_ub"), ("UM", "open_um"), ("UMK", "open_umk"), ("Keluarga", "open_kel")]
            n["sumber"] = "Progres Pendataan kol. (11) sampai (14)"
            n["turunan"] = ("Penjumlahan empat kolom Open di ekspor. Persentasenya tidak tersedia di berkas, "
                            "jadi yang ditampilkan adalah porsi terhadap total assignment.")

def jam(p):
    return datetime.datetime.fromtimestamp(p.stat().st_mtime).strftime("%d %b %Y %H.%M")

stempel = ("Sumber FASIH Pendataan SE2026<br>Pendataan " + jam(F1)
           + " &middot; Keluarga " + jam(F2)
           + "<br>Halaman disusun " + datetime.datetime.now().strftime("%d %b %Y %H.%M"))

muatan = json.dumps({"rows": rows, "kodes": kodes, "nodes": N, "stempel": stempel,
                     "catatan": CATATAN_EKSPOR},
                    ensure_ascii=False, separators=(",", ":"))

tpl = (AKAR / "template.html").read_text(encoding="utf-8")
if "__PAYLOAD__" not in tpl:
    sys.exit("GAGAL: template.html tidak memuat penanda __PAYLOAD__")
(AKAR / "index.html").write_text(tpl.replace("__PAYLOAD__", muatan), encoding="utf-8")

print(f"\nindex.html selesai ditulis ({len(rows)} wilayah, {len(N)} kartu).")
if not lulus or beda:
    print("PERINGATAN: ada uji yang tidak lulus. Periksa berkas ekspor sebelum dipublikasikan.")
